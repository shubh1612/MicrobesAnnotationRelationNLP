from nltk import sent_tokenize, word_tokenize
import bisect
import operator
import subprocess
import dependency_parse_tree as dtree
import openpyxl as pyxl
import re

threshold = 3

search_dest = 'trn/'

#make list of articles
subprocess.call('ls ../training_set/corrected_articles/'+search_dest+' | grep \'.txt\' > ../training_set/extra/file_list.txt' , shell = True)

with open('../training_set/extra/file_list.txt', 'r') as filel:
	file_names = filel.readlines()
file_names = [x.strip() for x in file_names]

###############################################################
class entity_obj:
	
	def __init__(self, num, type1, start, end, name):
		self.num = num
		self.type = type1
		self.start = start
		self.end = end
		self.name = name

	def string_entity(self):
		return str('T'+str(self.num)+'\t'+self.type+' '+str(self.start)+' '+str(self.end)+'\t'+self.name)

	def list_entity(self):
		l = []
		l.append(self.num)
		l.append(self.start)
		l.append(self.end)
		l.append(self.name)
		l.append(self)
		return l

###################################################################
class rel_obj:
	
	def __init__(self, num, type1, first, second):
		self.num = num
		self.type = type1
		self.first = first
		self.second = second

	def string_rel(self):
		return str(str(self.num)+'\t'+self.type+' Arg1:T'+str(self.first)+' Arg2:T'+str(self.second))	

	def contains(self, arg):
		if self.first == arg or self.second == arg:
			return 1
		else:
			return 0
	
	def remove_rel(self):
		self.first = -1

	def print_check(self):
		if self.first == -1 :
			return 0
		else:
			return 1
	

#######################################################################

removed_rel_list = []

for file in file_names:
	with open('../training_set/corrected_articles/'+search_dest+file[:-4]+'.ann', 'r') as myfile:
		data = myfile.readlines()
	data = [x.strip() for x in data]

	dict_entity = {}
	entity_obj_list = []
	rel_obj_list = []
	entity_list = []
	for line in data:
		if line[0] == 'T':
			word = word_tokenize(line)
			obj = entity_obj(int(word[0][1:]), word[1], int(word[2]), int(word[3]), ' '.join(word[4:]))
			dict_entity[int(word[0][1:])] = obj
			entity_obj_list.append(obj)
			entity_list.append(obj.list_entity())
		else:
			word = word_tokenize(line)
			obj = rel_obj(word[0], word[1], int(word[4][1:]), int(word[7][1:]))
			rel_obj_list.append(obj)
	

	with open('../training_set/corrected_articles/'+search_dest+file, 'r', encoding='utf-8') as myfile:
		data=myfile.read().replace('\n',' ')
	
	sort_entity_start = sorted(entity_list, key=operator.itemgetter(1,2))
	
	char_num = 0
	for sentence in sent_tokenize(data):
		print(sentence + '\n')
					
		sent_entity_start = bisect.bisect_left([x[1] for x in entity_list], char_num)
		sent_entity_end = bisect.bisect_right([x[1] for x in entity_list], char_num + len(sentence))

		sent_rel_list = []
		for x in rel_obj_list:
			for y in range(sent_entity_start, sent_entity_end):
				if y < len(entity_list):
					if x.contains(entity_list[y][0]) and x not in sent_rel_list:
						sent_rel_list.append(x)

		depen_tree = dtree.graph(sentence)
		adj_list = depen_tree[0]
		dict_char_begin = depen_tree[1] 
		list_char_begin = list(dict_char_begin.keys())
		dict_token = depen_tree[2]

		

		for rel in sent_rel_list:
			entity1 = dict_entity[rel.first]
			entity2 = dict_entity[rel.second]
			entity1_tokens = []
			entity2_tokens = []
			for char_st in list_char_begin:
				if char_st >= entity1.start - char_num and char_st < entity1.end - char_num:
					entity1_tokens.append(dict_char_begin[char_st])
				if char_st >= entity2.start - char_num and char_st < entity2.end - char_num:
					entity2_tokens.append(dict_char_begin[char_st])

			min_dist = len(sentence) + 1
			for i in entity1_tokens:
				for j in entity2_tokens:
					min_dist = min(dtree.dist_n1_n2(adj_list,i,j,len(dict_token)), min_dist)

			if min_dist > threshold:
				print(rel.num, rel.first, rel.second)
				temp_list = [file , sentence, dict_entity[rel.first], dict_entity[rel.second], min_dist]
				removed_rel_list.append(temp_list)
				rel.remove_rel()

		char_num += len(sentence) + 1
		if len(sentence) == 1:
			char_num += 1
	
	fp = open('../training_set/corrected_articles/'+search_dest+file[:-4]+'.ann', 'w', encoding='utf-8')
	for i in entity_obj_list:
		fp.write(i.string_entity() + '\n')
	for i in rel_obj_list:
		if i.print_check() == 1:
			fp.write(i.string_rel() + '\n')
	fp.close() 
	print(file)

#######################################################################################333

wb = pyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')

r = 1
c = 1

sheet.cell(row = r, column = c).value = 'file name'
sheet.cell(row = r, column = c+1).value = 'sentence'
sheet.cell(row = r, column = c+2).value = 'first_entity'
sheet.cell(row = r, column = c+3).value = 'second_entity'
sheet.cell(row = r, column = c+4).value =  'distance in dependency_parse_tree'
	
r += 2		
prev = removed_rel_list[0][0]
for x in removed_rel_list:
	if x[0] != prev:
		r += 1 
	sheet.cell(row = r, column = c).value = x[0]
	sheet.cell(row = r, column = c+1).value = x[1]
	sheet.cell(row = r, column = c+2).value = x[2].name
	sheet.cell(row = r, column = c+3).value = x[3].name
	sheet.cell(row = r, column = c+4).value = x[4]
	r += 1
	prev = x[0]

wb.save('result_sent.xlsx')