import numpy as np 
from nltk import word_tokenize
import subprocess
import openpyxl as pyxl
import os 


relation_map = {"FoundIn": 0, "Of": 1, "ActOn": 2, "CarriedOn": 3, "Produce": 4, "OccursIn": 5, "Inhibit_Catalyse": 6}

# subprocess.call('ls org_articles/ | grep \'.ann\' > file_list.txt')
with open('file_list_1.txt', 'r') as inp_file:
	files = inp_file.readlines()
files = [x.strip() for x in files]

class EntityObj:
	def __init__(self, num, type_inp, start, end, name):
		self.num = num
		self.type = type_inp
		self.start = start
		self.end = end
		self.name = name

class RelObj:
	def __init__(self, num, type_inp, first, second):
		self.num = num
		self.type = type_inp
		self.first = first
		self.second = second

true_positive = [0]*7
false_negative = [0]*7
total_new = [0]*7

tempc = 0
for file in files:
	with open(file, 'r') as input_file:
			org_ann = input_file.readlines()
	org_ann = [x.strip() for x in org_ann]

	entity_map = {}
	for line in org_ann:
		if line[0] == 'T':
			word = word_tokenize(line)
			if len(word) < 5 :
				continue
			obj = EntityObj(int(word[0][1:]), word[1], int(word[2]), int(word[3]), ' '.join(word[4:]))
			entity_map[obj.num] = obj
	
	org_relation_list = []
	for line in org_ann:
		if line[0] == 'R':
			word = word_tokenize(line)
			if len(word) < 8 :
				continue
			if word[1][-3:] == '_No':
				continue
			obj = RelObj(int(word[0][1:]), word[1], int(word[4][1:]), int(word[7][1:]))
			if obj.first not in entity_map.keys() or obj.second not in entity_map.keys():
				continue
			tp = (obj.type, (entity_map[obj.first].start, entity_map[obj.first].end), (entity_map[obj.second].start, entity_map[obj.second].end))
			org_relation_list.append(tp)

	if os.path.exists(file[0:-4] + '__.ann') == False:
		print(file + '\n')
		continue

	with open(file[0:-4] + '__.ann', 'r') as input_file:
			new_ann = input_file.readlines()
	new_ann = [x.strip() for x in new_ann]

	entity_map = {}
	for line in new_ann:
		if line[0] == 'T':
			word = word_tokenize(line)
			if len(word) < 5 :
				continue
			obj = EntityObj(int(word[0][1:]), word[1], int(word[2]), int(word[3]), ' '.join(word[4:]))
			entity_map[obj.num] = obj
	
	new_relation_list = []
	for line in new_ann:
		if line[0] == 'R':
			word = word_tokenize(line)
			if len(word) < 8 :
				continue
			if word[1][-3:] == '_No':
				continue
			obj = RelObj(int(word[0][1:]), word[1], int(word[4][1:]), int(word[7][1:]))
			if obj.type == 'CarriedOn':
				print(file)
			if 'Carried' in obj.type :
				print(obj.type, file)
			if obj.first not in entity_map.keys() or obj.second not in entity_map.keys():
				continue
			tp = (obj.type, (entity_map[obj.first].start, entity_map[obj.first].end), (entity_map[obj.second].start, entity_map[obj.second].end))
			new_relation_list.append(tp)
	
	for new_rel in new_relation_list:
		total_new[relation_map[new_rel[0]]] += 1
		for org_rel in org_relation_list:
			if org_rel == new_rel:
				true_positive[relation_map[new_rel[0]]] += 1

	for org_rel in org_relation_list:
		if org_rel not in new_relation_list:
			false_negative[relation_map[org_rel[0]]] += 1

	# print(file)
	for org_rel in org_relation_list:
		if org_rel[0] == 'CarriedOn':
			tempc += 1
print(tempc)

wb = pyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')

c = 1
r = 1
sheet.cell(row = r, column = c).value = 'Relation' 


c += 2

sheet.cell(row = r, column = c).value = 'True Positive' 
sheet.cell(row = r, column = c+1).value = 'False Positive'
sheet.cell(row = r, column = c+2).value = 'False Negative'

sheet.cell(row = r, column = c+3).value = 'Precision ( tp/(tp + fp) )' 
sheet.cell(row = r, column = c+4).value = 'Recall ( tp/(tp + fn) )'
sheet.cell(row = r, column = c+5).value = 'F1 Score'

r += 2

for relation in relation_map.keys():
	c = 1
	sheet.cell(row = r, column = c).value = relation

	c += 2

	tp = true_positive[relation_map[relation]]
	fp = total_new[relation_map[relation]] - true_positive[relation_map[relation]]
	fn = false_negative[relation_map[relation]]	
	if (tp + fp) == 0:
		prec = np.inf
	else:
		prec = tp/(tp + fp)
	if (tp + fn) == 0:
		rec = np.inf
	else:
		rec = tp/(tp + fn)
	if (prec + rec) == 0:
		f1 = np.inf
	else:
		f1 = 2*prec*rec/(prec + rec)

	sheet.cell(row = r, column = c).value = tp
	sheet.cell(row = r, column = c+1).value = fp
	sheet.cell(row = r, column = c+2).value = fn

	sheet.cell(row = r, column = c+3).value = prec
	sheet.cell(row = r, column = c+4).value = rec
	sheet.cell(row = r, column = c+5).value = f1

	r += 2

wb.save('relations_f1.xlsx')
