import subprocess
import openpyxl as pyxl
from nltk import word_tokenize


class EntityObj:

	def __init__(self, num, type_inp, start, end, name):
		self.num = num
		self.type = type_inp
		self.start = start
		self.end = end
		self.name = name
	
	def list_entity(self):
		l = []
		l.append(self.num)
		l.append(self.start)
		l.append(self.end)
		l.append(self)
		return l


class RelObj:
	
	def __init__(self, num, type_inp, first, second):
		self.num = num
		self.type = type_inp
		self.first = first
		self.second = second

	def contains(self, arg):
		if self.first == arg or self.second == arg:
			return 1
		else:
			return 0	


def word_probability():

	global relation_map, probability_relation_list

	relation_map = {"FoundIn": '0', "FoundIn_No": '1', "Of": '2', "Of_No": '3', "ActOn": '4', "ActOn_No": '5',
					"CarriedOn": '6',"CarriedOn_No": '7', "Produce": '8', "Produce_No": '9', "OccursIn": '10',
					"OccursIn_No": '11', "Inhibit_Catalyse": '12', "Inhibit_Catalyse_No": '13'}

	probability_relation_list = [{} for i in range(len(relation_map))]

	for relation in relation_map.keys():
		dict = {}
		
		with open('files_pos_neg/' + relation_map[relation] + '.txt', 'r') as file:
			data = file.readlines()
		data = [x.strip() for x in data]
		data = [x.lower() for x in data]
		n = len(data)
		
		for line in data:
			for word in word_tokenize(line):
				if word in dict:
					dict[word] = dict[word] + 1/n
				else:
					dict[word] = 1/n

		probability_relation_list[int(relation_map[relation])] = dict


def calculate_probability():

	global probability_list
	probability_list = [[] for i in range(len(relation_map))]

	subprocess.call('ls ' + search_dest + ' | grep \'.txt\' > ../../training_set/extra/file_list.txt', shell = True)

	with open('../training_set/extra/file_list.txt', 'r') as input_file:
		file_list = input_file.readlines()
	file_list = [x.strip() for x in file_list]

	for file in file_list:
		with open(search_dest + file, 'r') as input_file:
			data=input_file.read().replace('\n',' ')
		
		with open(search_dest + file[:-4] + '.ann', 'r') as input_file:
			ann_input=input_file.readlines()
		ann_input = [x.strip() for x in ann_input]

		entity_map = {}
		entity_obj_list = []
		relation_list = []
		for line in ann_input:
			if line[0] == 'T':
				word = word_tokenize(line)
				# print(word)
				if len(word) < 5 :
					continue
				obj = EntityObj(int(word[0][1:]), word[1], int(word[2]), int(word[3]), ' '.join(word[4:]))
				entity_obj_list.append(obj)
				entity_map[obj.num] = obj
			else:
				word = word_tokenize(line)
				if len(word) < 8 :
					continue
				obj = RelObj(int(word[0][1:]), word[1], int(word[4][1:]), int(word[7][1:]))
				relation_list.append(obj)

		for relation in relation_list:
			rel_num_even = (int(relation_map[relation.type]) >> 1) << 1
			probability = 0
			if relation.first not in entity_map.keys() or relation.second not in entity_map.keys():
				continue
			start = min(entity_map[relation.first].end, entity_map[relation.second].end)
			end = max(entity_map[relation.first].start, entity_map[relation.second].start)
			line = data[start:end] 
			for word in word_tokenize(line):
				if word in probability_relation_list[rel_num_even]:
					probability += probability_relation_list[rel_num_even][word]
				if word in probability_relation_list[rel_num_even + 1]:
					probability -= probability_relation_list[rel_num_even + 1][word]
			
			rel_tuple = [file, '<' + entity_map[relation.first].name + ', ' + relation.type + ', ' + entity_map[relation.second].name + '>', probability]

			probability_list[int(relation_map[relation.type])].append(rel_tuple)
		
		print(file)		
	
def data_excel():
	wb = pyxl.Workbook()
	sheet = wb.get_sheet_by_name('Sheet')

	c = 1

	for relation in relation_map.keys():
		r = 1	

		sheet.cell(row = r, column = c).value = relation
		r += 3

		sheet.cell(row = r, column = c).value = 'File' 
		sheet.cell(row = r, column = c+1).value = 'Relation'
		sheet.cell(row = r, column = c+2).value = 'Probability'

		r += 2
		
		for rel_tuple in probability_list[int(relation_map[relation])]:
			sheet.cell(row = r, column = c).value = rel_tuple[0] 
			sheet.cell(row = r, column = c+1).value = rel_tuple[1]
			sheet.cell(row = r, column = c+2).value = rel_tuple[2]
			r += 1
		
		c += 5 

	wb.save('relations_probability.xlsx')



def main():
	global search_dest
	search_dest = '../../training_set/corrected_articles/biotech_corrected/'

	word_probability()
	calculate_probability()
	data_excel()

main()