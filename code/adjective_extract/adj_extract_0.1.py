#execute in the stanford-nlp directory
#author - manas 

import subprocess
import csv
import openpyxl as pyxl

#put all articles in directory articles
#make list of all articles 
subprocess.call('ls articles\\ > extra\\file_list.txt' , shell = True)

with open('extra\\file_list.txt', 'r') as filel:
	file_names = filel.readlines()
file_names = [x.strip() for x in file_names]


###########make general list.txt and import
with open('extra\\entity_list.txt', 'r') as f_l:
	entity = f_l.readlines()
entity = [x.strip() for x in entity]
# entity = ['abiotic', 'aerobic','aliphatic','anoxic','anaerobic','anthropogenic','aquatic','biodegradation','bioremediation','catabolic','decontamination','degradation','denaturing','denitrifying','detoxification','eukaryotic','fermentation','fertilization','ferric','heterotrophic','humic','methanogenic','mineralization','oligotrophic','prokaryotic','syntrophic','thermophilic','toxicity','xenobiotic','aromatic','antarctic','beach','brazil','sea','gasoline','groundwater',	'hydrocarbon','nutrient','landfill','lake',	'marsh','oil','petroleum','sand','seawater','tree','wastewater','sediment','site','soil','communities']
extra_adj = []
dict = {}
for obj in entity:
	l = []
	dict[obj] = l

# import stopwords from stopword.txt in the extra directory 
with open('extra\stopwords.txt', 'r', encoding = 'utf-8') as f1:
	stopword = f1.readlines()
stopword = [x.strip() for x in stopword]

# import manual stopwords from man_stopword.txt in the extra directory
with open('extra\man_stopwords.txt', 'r', encoding = 'utf-8') as f2:
	stopword = f2.readlines()
stopword = [x.strip() for x in stopword]

num_s = 6

def mk_l(j, fl):
	l = [word[j], ' '.join(word[j-num_s : j+num_s]), fl]
	return l

for fl in file_names:
	# java command - note the name and destination of file 
	subprocess.call('java -cp "*" -Xmx1g edu.stanford.nlp.pipeline.StanfordCoreNLP -annotators tokenize,ssplit,pos,lemma -file articles\\'+ fl, shell = True)
	
	# move .xml file to a different directory
	subprocess.call('mv ' + fl + '.xml article_xmls\\', shell = True)	
	
	with open('article_xmls\\' + fl +'.xml',  'r', encoding = 'utf-8') as file:
		I = file.readlines()
	I = [x.strip() for x in I] 

	#make pos and array lists
	word = []
	pos = []
	for line in I:
		if '<word>' in line and '</word>' in line:
			word.append(line[6:-7].lower())
		if '<POS>' in line and '</POS>' in line:
			pos.append(line[5:-6])

	for obj in entity:
		for i in range(1,len(word)):
			if word[i] == obj or word[i] == obj+'s':
				j = i-1
				x = obj
				while pos[j] == 'JJ':
					if x not in dict.keys():
						l = []
						dict[x] = l
					if word[j] in stopword:
						break;
					if mk_l(j,fl) not in dict[x]:
						dict[x].append(mk_l(j,fl))
						if '-' in word[j]:
							extra_adj.append(word[j][word[j].find('-')+1:])
					x = word[j] + ' ' +obj
					j -= 1
	for obj in entity:
		for i in range(1,len(word)):
			if word[i] == obj  or word[i] == obj+'s':
				j = i-1
				x = obj
				if word[j] in extra_adj:
					if word[j][-3:] == 'ing':
						if mk_l(j,fl) in dict[x]:
							dict[x].remove(mk_l(j,fl))
						if [word[j-1] + ' ' + word[j],' '.join(word[j-num_s : j+num_s]), fl] not in dict[x]:
							dict[x].append([word[j-1] + ' ' + word[j],' '.join(word[j-num_s : j+num_s]), fl])
					else:
						if mk_l(j,fl) not in dict[x]:
							if pos[j-1] == 'NN' :
								dict[x].append([word[j-1] + ' ' + word[j],' '.join(word[j-num_s : j+num_s]), fl])	 
							else:
								dict[x].append(mk_l(j,fl))
						
	print(fl)
	#loop end
print(set(extra_adj))

# convert into Excel file
wb = pyxl.Workbook()
sheet = wb.get_sheet_by_name('Sheet')

r = 1
c = 1
for x in dict.keys():
	r = 1
	sheet.cell(row = r, column = c).value = x

	r = 3
	if len(dict[x])!=0:
		for y in dict[x]:
			sheet.cell(row = r, column = c).value = y[0]
			sheet.cell(row = r, column = c+1).value = y[1]
			sheet.cell(row = r, column = c+2).value = y[2]
			sheet.cell(row = r, column = c+3).value = '.'
			r += 1
		c += 6;

wb.save('result_sent.xlsx')
